import streamlit as st
import os
from openpyxl import Workbook, load_workbook

FILE_NAME = "natak_bookings.xlsx"

# Create Excel file if not exists
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Mobile", "Seat", "Date", "Time"])
    wb.save(FILE_NAME)

st.title("🎭 Natak Online Booking")

name = st.text_input("Enter Name")
mobile = st.text_input("Mobile Number")

seat = st.selectbox(
    "Select Seat",
    ["A1","A2","A3","A4","B1","B2","B3","B4"]
)

show_date = st.date_input("Show Date")

show_time = st.selectbox(
    "Show Time",
    ["4:00 PM", "7:00 PM"]
)

if st.button("Book Seat"):

    if name and mobile:

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        ws.append([
            name,
            mobile,
            seat,
            str(show_date),
            show_time
        ])

        wb.save(FILE_NAME)

        st.success("✅ Booking Saved to Excel!")

    else:
        st.error("Please fill all details")